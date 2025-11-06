# ========= Copyright 2023-2024 @ CAMEL-AI.org. All Rights Reserved. =========
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ========= Copyright 2023-2024 @ CAMEL-AI.org. All Rights Reserved. =========

# Enables postponed evaluation of annotations (for string-based type hints)
import os
from pathlib import Path
from typing import TYPE_CHECKING, List, Optional, Union

from camel.logger import get_logger
from camel.toolkits.base import BaseToolkit
from camel.toolkits.function_tool import FunctionTool
from camel.utils import MCPServer

# Import only for type hints (not executed at runtime)
if TYPE_CHECKING:
    from pandas import DataFrame

logger = get_logger(__name__)


@MCPServer()
class ExcelToolkit(BaseToolkit):
    r"""A class representing a toolkit for extract detailed cell information
    from an Excel file.

    This class provides methods extracting detailed content from Excel files
    (including .xls, .xlsx,.csv), and converting the data into
    Markdown formatted table.
    """

    def __init__(
        self,
        timeout: Optional[float] = None,
        working_directory: Optional[str] = None,
    ):
        r"""Initializes a new instance of the ExcelToolkit class.

        Args:
            timeout (Optional[float]): The timeout value for API requests
                in seconds. If None, no timeout is applied.
                (default: :obj:`None`)
            working_directory (str, optional): The default directory for
                output files. If not provided, it will be determined by the
                `CAMEL_WORKDIR` environment variable (if set). If the
                environment variable is not set, it defaults to
                `camel_working_dir`.
        """
        super().__init__(timeout=timeout)
        self.wb = None
        if working_directory:
            self.working_directory = Path(working_directory).resolve()
        else:
            camel_workdir = os.environ.get("CAMEL_WORKDIR")
            if camel_workdir:
                self.working_directory = Path(camel_workdir).resolve()
            else:
                self.working_directory = Path("./camel_working_dir").resolve()

        self.working_directory.mkdir(parents=True, exist_ok=True)
        logger.info(
            f"ExcelToolkit initialized with output directory: "
            f"{self.working_directory}"
        )

    def _validate_file_path(self, file_path: str) -> bool:
        r"""Validate file path for security.

        Args:
            file_path (str): The file path to validate.

        Returns:
            bool: True if path is safe, False otherwise.
        """
        normalized_path = os.path.normpath(file_path)

        if '..' in normalized_path.split(os.path.sep):
            return False

        return True

    def _convert_to_markdown(self, df: "DataFrame") -> str:
        r"""Convert DataFrame to Markdown format table.

        Args:
            df (DataFrame): DataFrame containing the Excel data.

        Returns:
            str: Markdown formatted table.
        """
        from tabulate import tabulate

        md_table = tabulate(df, headers='keys', tablefmt='pipe')
        return str(md_table)

    def extract_excel_content(self, document_path: str) -> str:
        r"""Extract and analyze the full content of an Excel file (.xlsx/.xls/.
        csv).

        Use this tool to read and understand the structure and content of
        Excel files. This is typically the first step when working with
        existing Excel files.

        Args:
            document_path (str): The file path to the Excel file.

        Returns:
            str: A comprehensive report containing:
                - Sheet names and their content in markdown table format
                - Detailed cell information including values, colors, and
                    positions
                - Formatted data that's easy to understand and analyze
        """
        import pandas as pd
        from openpyxl import load_workbook
        from xls2xlsx import XLS2XLSX

        logger.debug(
            f"Calling extract_excel_content with document_path"
            f": {document_path}"
        )

        if not self._validate_file_path(document_path):
            return "Error: Invalid file path."

        if not (
            document_path.endswith("xls")
            or document_path.endswith("xlsx")
            or document_path.endswith("csv")
        ):
            logger.error("Only xls, xlsx, csv files are supported.")
            return (
                f"Failed to process file {document_path}: "
                f"It is not excel format. Please try other ways."
            )

        if not os.path.exists(document_path):
            return f"Error: File {document_path} does not exist."

        if document_path.endswith("csv"):
            try:
                df = pd.read_csv(document_path)
                md_table = self._convert_to_markdown(df)
                return f"CSV File Processed:\n{md_table}"
            except Exception as e:
                logger.error(f"Failed to process file {document_path}: {e}")
                return f"Failed to process file {document_path}: {e}"

        if document_path.endswith("xls"):
            output_path = document_path.replace(".xls", ".xlsx")
            x2x = XLS2XLSX(document_path)
            x2x.to_xlsx(output_path)
            document_path = output_path

        try:
            # Load the Excel workbook
            wb = load_workbook(document_path, data_only=True)
            sheet_info_list = []

            # Iterate through all sheets
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                cell_info_list = []

                for row in ws.iter_rows():
                    for cell in row:
                        # Skip cells that don't have proper coordinates (like
                        # merged cells)
                        if (
                            not hasattr(cell, 'column_letter')
                            or cell.value is None
                        ):
                            continue

                        row_num = cell.row
                        # Use getattr with fallback for column_letter
                        col_letter = getattr(cell, 'column_letter', 'A')

                        cell_value = cell.value

                        font_color = None
                        if (
                            cell.font
                            and cell.font.color
                            and "rgb=None" not in str(cell.font.color)
                        ):  # Handle font color
                            font_color = cell.font.color.rgb

                        fill_color = None
                        if (
                            cell.fill
                            and cell.fill.fgColor
                            and "rgb=None" not in str(cell.fill.fgColor)
                        ):  # Handle fill color
                            fill_color = cell.fill.fgColor.rgb

                        cell_info_list.append(
                            {
                                "index": f"{row_num}{col_letter}",
                                "value": cell_value,
                                "font_color": font_color,
                                "fill_color": fill_color,
                            }
                        )

                # Convert the sheet to a DataFrame and then to markdown
                sheet_df = pd.read_excel(
                    document_path, sheet_name=sheet, engine='openpyxl'
                )
                markdown_content = self._convert_to_markdown(sheet_df)

                # Collect all information for the sheet
                sheet_info = {
                    "sheet_name": sheet,
                    "cell_info_list": cell_info_list,
                    "markdown_content": markdown_content,
                }
                sheet_info_list.append(sheet_info)

            result_str = ""
            for sheet_info in sheet_info_list:
                result_str += f"""
            Sheet Name: {sheet_info['sheet_name']}
            Cell information list:
            {sheet_info['cell_info_list']}
            
            Markdown View of the content:
            {sheet_info['markdown_content']}
            
            {'-'*40}
            """

            return result_str
        except Exception as e:
            logger.error(f"Failed to process Excel file {document_path}: {e}")
            return f"Failed to process Excel file {document_path}: {e}"

    def _save_workbook(self, file_path: str) -> str:
        r"""Save the current workbook to file.

        Args:
            file_path (str): The path to save the workbook.

        Returns:
            str: Success or error message.
        """
        if not self.wb:
            return "Error: No workbook loaded to save."

        if not self._validate_file_path(file_path):
            return "Error: Invalid file path for saving."

        try:
            self.wb.save(file_path)
            return f"Workbook saved successfully to {file_path}"
        except Exception as e:
            logger.error(f"Failed to save workbook: {e}")
            return f"Error: Failed to save workbook: {e}"

    def save_workbook(self, filename: str) -> str:
        r"""Save the current in-memory workbook to a file.

        Args:
            filename (str): The filename to save the workbook. Must end with
                .xlsx extension. The file will be saved in self.
                working_directory.

        Returns:
            str: Success message or error details.
        """
        if not self.wb:
            return "Error: No workbook is currently loaded in memory."

        # Validate filename
        if not filename:
            return "Error: Filename is required."

        if not filename.endswith('.xlsx'):
            return "Error: Filename must end with .xlsx extension."

        # Create full path in working directory
        file_path = self.working_directory / filename
        resolved_file_path = str(file_path.resolve())

        return self._save_workbook(resolved_file_path)

    def create_workbook(
        self,
        filename: Optional[str] = None,
        sheet_name: Optional[str] = None,
        data: Optional[List[List[Union[str, int, float, None]]]] = None,
    ) -> str:
        r"""Create a new Excel workbook from scratch.

        Use this when you need to create a new Excel file. This sets up the
        toolkit to work with the new file and optionally adds initial data.

        Args:
            filename (Optional[str]): The filename for the workbook. Must end
                with .xlsx extension. The file will be saved in
                self.working_directory. (default: :obj:`None`)
            sheet_name (Optional[str]): Name for the first sheet. If None,
                creates "Sheet1". (default: :obj:`None`)
            data (Optional[List[List[Union[str, int, float, None]]]]): Initial
                data as rows. Each inner list is one row. (default:
                :obj:`None`)

        Returns:
            str: Success confirmation message or error details
        """
        from openpyxl import Workbook

        # Validate filename
        if filename is None:
            return "Error: Filename is required."

        if not filename.endswith('.xlsx'):
            return "Error: Filename must end with .xlsx extension."

        # Create full path in working directory
        file_path = self.working_directory / filename
        resolved_file_path = str(file_path.resolve())

        if not self._validate_file_path(resolved_file_path):
            return "Error: Invalid file path."

        # Check if file already exists
        if os.path.exists(resolved_file_path):
            return (
                f"Error: File {filename} already exists in "
                f"{self.working_directory}."
            )

        try:
            # Create a new workbook
            wb = Workbook()
            self.wb = wb

            # Handle sheet creation safely
            if sheet_name:
                # Remove the default sheet safely
                default_sheet = wb.active
                if default_sheet is not None:
                    wb.remove(default_sheet)
                ws = wb.create_sheet(sheet_name)
            else:
                ws = wb.active
                if ws is not None and sheet_name is None:
                    sheet_name = "Sheet1"
                    ws.title = sheet_name

            # Add data if provided
            if data and ws is not None:
                for row in data:
                    ws.append(row)

            # Save the workbook to the specified file path
            wb.save(resolved_file_path)

            return f"Workbook created successfully at {resolved_file_path}"
        except Exception as e:
            logger.error(f"Failed to create workbook: {e}")
            return f"Error: Failed to create workbook: {e}"

    def delete_workbook(self, filename: str) -> str:
        r"""Delete a spreadsheet file from the working directory.

        Args:
            filename (str): The filename to delete. Must end with .xlsx
                extension. The file will be deleted from self.
                working_directory.

        Returns:
            str: Success message or error details.
        """
        # Validate filename
        if not filename:
            return "Error: Filename is required."

        if not filename.endswith('.xlsx'):
            return "Error: Filename must end with .xlsx extension."

        # Create full path in working directory
        file_path = self.working_directory / filename
        target_path = str(file_path.resolve())

        if not self._validate_file_path(target_path):
            return "Error: Invalid file path."

        if not os.path.exists(target_path):
            return (
                f"Error: File {filename} does not exist in "
                f"{self.working_directory}."
            )

        try:
            os.remove(target_path)
            # Clean up workbook if one is loaded
            self.wb = None
            return (
                f"Workbook {filename} deleted successfully from "
                f"{self.working_directory}."
            )
        except Exception as e:
            logger.error(f"Failed to delete workbook: {e}")
            return f"Error: Failed to delete workbook {filename}: {e}"

    def create_sheet(
        self,
        sheet_name: str,
        data: Optional[List[List[Union[str, int, float, None]]]] = None,
    ) -> str:
        r"""Create a new sheet with the given sheet name and data.

        Args:
            sheet_name (str): The name of the sheet to create.
            data (Optional[List[List[Union[str, int, float, None]]]]):
                The data to write to the sheet.

        Returns:
            str: Success message.
        """
        if not self.wb:
            return (
                "Error: Workbook not initialized. "
                "Please create a workbook first."
            )

        if sheet_name in self.wb.sheetnames:
            return f"Error: Sheet {sheet_name} already exists."

        try:
            ws = self.wb.create_sheet(sheet_name)
            if data:
                for row in data:
                    ws.append(row)

            return f"Sheet {sheet_name} created successfully."
        except Exception as e:
            logger.error(f"Failed to create sheet: {e}")
            return f"Error: Failed to create sheet {sheet_name}: {e}"

    def delete_sheet(self, sheet_name: str) -> str:
        r"""Delete a sheet from the workbook.

        Args:
            sheet_name (str): The name of the sheet to delete.

        Returns:
            str: Success message.
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Sheet {sheet_name} does not exist."

        if len(self.wb.sheetnames) == 1:
            return "Cannot delete the last remaining sheet in the workbook."

        try:
            ws = self.wb[sheet_name]
            self.wb.remove(ws)
            return f"Sheet {sheet_name} deleted successfully."
        except Exception as e:
            logger.error(f"Failed to delete sheet: {e}")
            return f"Error: Failed to delete sheet {sheet_name}: {e}"

    def clear_sheet(self, sheet_name: str) -> str:
        r"""Clear all data from a sheet.

        Args:
            sheet_name (str): The name of the sheet to clear.

        Returns:
            str: Success message.
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Sheet {sheet_name} does not exist."

        try:
            ws = self.wb[sheet_name]

            # Clear all cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None

            return f"Sheet {sheet_name} cleared successfully."
        except Exception as e:
            logger.error(f"Failed to clear sheet: {e}")
            return f"Error: Failed to clear sheet {sheet_name}: {e}"

    def delete_rows(
        self, sheet_name: str, start_row: int, end_row: Optional[int] = None
    ) -> str:
        r"""Delete rows from a sheet.

        Use this to remove unwanted rows. You can delete single rows or ranges.

        Args:
            sheet_name (str): Name of the sheet to modify.
            start_row (int): Starting row number to delete (1-based, where 1
                is first row).
            end_row (Optional[int]): Ending row number to delete (1-based).
                If None, deletes only start_row. (default: :obj:`None`)

        Returns:
            str: Success confirmation message or error details
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Sheet {sheet_name} does not exist."

        ws = self.wb[sheet_name]

        if end_row is None:
            end_row = start_row

        # Delete rows (openpyxl uses 1-based indexing)
        num_rows = end_row - start_row + 1
        ws.delete_rows(start_row, num_rows)

        return (
            f"Deleted rows {start_row} to {end_row} from sheet "
            f"{sheet_name} successfully."
        )

    def delete_columns(
        self, sheet_name: str, start_col: int, end_col: Optional[int] = None
    ) -> str:
        r"""Delete columns from a sheet.

        Use this to remove unwanted columns. You can delete single columns or
        ranges.

        Args:
            sheet_name (str): Name of the sheet to modify.
            start_col (int): Starting column number to delete (1-based, where
                1 is column A).
            end_col (Optional[int]): Ending column number to delete (1-based).
                If None, deletes only start_col. (default: :obj:`None`)

        Returns:
            str: Success confirmation message or error details
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Sheet {sheet_name} does not exist."

        ws = self.wb[sheet_name]

        if end_col is None:
            end_col = start_col

        # Delete columns (openpyxl uses 1-based indexing)
        num_cols = end_col - start_col + 1
        ws.delete_cols(start_col, num_cols)

        return (
            f"Deleted columns {start_col} to {end_col} from sheet "
            f"{sheet_name} successfully."
        )

    def get_cell_value(
        self, sheet_name: str, cell_reference: str
    ) -> Union[str, int, float, None]:
        r"""Get the value from a specific cell.

        Use this to read a single cell's value. Useful for checking specific
        data points or getting values for calculations.

        Args:
            sheet_name (str): Name of the sheet containing the cell.
            cell_reference (str): Excel-style cell reference (column letter +
                row number).

        Returns:
            Union[str, int, float, None]: The cell's value or error message
                Returns None for empty cells.
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Error: Sheet {sheet_name} does not exist."

        ws = self.wb[sheet_name]
        return ws[cell_reference].value

    def set_cell_value(
        self,
        sheet_name: str,
        cell_reference: str,
        value: Union[str, int, float, None],
    ) -> str:
        r"""Set the value of a specific cell.

        Use this to update individual cells with new values. Useful for
        corrections, calculations, or updating specific data points.

        Args:
            sheet_name (str): Name of the sheet containing the cell.
            cell_reference (str): Excel-style cell reference (column letter +
                row number).
            value (Union[str, int, float, None]): New value for the cell.
                (default: :obj:`None`)

        Returns:
            str: Success confirmation message or error details.
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Sheet {sheet_name} does not exist."

        try:
            ws = self.wb[sheet_name]
            # Handle None values properly - openpyxl doesn't accept None
            # directly
            if value is None:
                ws[cell_reference].value = None
            else:
                ws[cell_reference] = value
            return (
                f"Cell {cell_reference} updated successfully in sheet "
                f"{sheet_name}."
            )
        except Exception as e:
            logger.error(f"Failed to set cell value: {e}")
            return f"Error: Failed to set cell value: {e}"

    def get_column_data(
        self, sheet_name: str, column: Union[int, str]
    ) -> Union[List[Union[str, int, float, None]], str]:
        r"""Get all data from a specific column.

        Use this to extract all values from a column for analysis or
        processing.

        Args:
            sheet_name (str): Name of the sheet to read from.
            column (Union[int, str]): Column identifier - either number
                (1-based) or letter.

        Returns:
            Union[List[Union[str, int, float, None]], str]:
                List of all non-empty values in the column or error message
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Error: Sheet {sheet_name} does not exist."

        ws = self.wb[sheet_name]

        if isinstance(column, str):
            col_letter = column.upper()
        else:
            from openpyxl.utils import (  # type: ignore[import]
                get_column_letter,
            )

            col_letter = get_column_letter(column)

        column_data = []
        for cell in ws[col_letter]:
            if cell.value is not None:
                column_data.append(cell.value)

        return column_data

    def find_cells(
        self,
        sheet_name: str,
        search_value: Union[str, int, float],
        search_column: Optional[Union[int, str]] = None,
    ) -> Union[List[str], str]:
        r"""Find cells containing a specific value.

        Use this to locate where specific data appears in the sheet.

        Args:
            sheet_name (str): Name of the sheet to search in.
            search_value (Union[str, int, float]): Value to search for.
            search_column (Optional[Union[int, str]]): Limit search to
                specific column. If None, searches entire sheet. (default:
                :obj:`None`)

        Returns:
            Union[List[str], str]: List of cell references (like "A5", "B12")
                where the value was found, or error message.
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Error: Sheet {sheet_name} does not exist."

        ws = self.wb[sheet_name]
        found_cells = []

        if search_column:
            # Search in specific column
            if isinstance(search_column, str):
                col_letter = search_column.upper()
                for cell in ws[col_letter]:
                    if cell.value == search_value:
                        found_cells.append(cell.coordinate)
            else:
                from openpyxl.utils import (  # type: ignore[import]
                    get_column_letter,
                )

                col_letter = get_column_letter(search_column)
                for cell in ws[col_letter]:
                    if cell.value == search_value:
                        found_cells.append(cell.coordinate)
        else:
            # Search entire sheet
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value == search_value:
                        found_cells.append(cell.coordinate)

        return found_cells

    def get_range_values(
        self, sheet_name: str, cell_range: str
    ) -> Union[List[List[Union[str, int, float, None]]], str]:
        r"""Get values from a specific range of cells.

        Use this to read a rectangular block of cells at once.

        Args:
            sheet_name (str): Name of the sheet to read from.
            cell_range (str): Range in Excel format (start:end).

        Returns:
            Union[List[List[Union[str, int, float, None]]], str]:
                2D list where each inner list is a row of cell values, or
                error message.
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Error: Sheet {sheet_name} does not exist."

        ws = self.wb[sheet_name]
        range_values = []

        for row in ws[cell_range]:
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            range_values.append(row_values)

        return range_values

    def set_range_values(
        self,
        sheet_name: str,
        cell_range: str,
        values: List[List[Union[str, int, float, None]]],
    ) -> str:
        r"""Set values for a specific range of cells.

        Use this to update multiple cells at once with a 2D array of data.

        Args:
            sheet_name (str): Name of the sheet to modify.
            cell_range (str): Range in Excel format to update.
            values (List[List[Union[str, int, float, None]]]): 2D array of
                values. Each inner list represents a row.

        Returns:
            str: Success confirmation message or error details.
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Error: Sheet {sheet_name} does not exist."

        ws = self.wb[sheet_name]

        # Get the range
        cell_range_obj = ws[cell_range]

        # If it's a single row or column, convert to 2D format
        if not isinstance(cell_range_obj[0], tuple):
            cell_range_obj = [cell_range_obj]

        for row_idx, row in enumerate(cell_range_obj):
            if row_idx < len(values):
                for col_idx, cell in enumerate(row):
                    if col_idx < len(values[row_idx]):
                        cell.value = values[row_idx][col_idx]

        return f"Values set for range {cell_range} in sheet {sheet_name}."

    def export_sheet_to_csv(self, sheet_name: str, csv_filename: str) -> str:
        r"""Export a specific sheet to CSV format.

        Use this to convert Excel sheets to CSV files for compatibility or
        data exchange.

        Args:
            sheet_name (str): Name of the sheet to export.
            csv_filename (str): Filename for the CSV file. Must end with .csv
                extension. The file will be saved in self.working_directory.

        Returns:
            str: Success confirmation message or error details.
        """
        if not self.wb:
            return (
                "Error: No workbook is currently loaded. Use "
                "extract_excel_content to load a workbook first."
            )

        if sheet_name not in self.wb.sheetnames:
            return (
                f"Error: Sheet {sheet_name} does not exist in the current "
                "workbook."
            )

        # Validate filename
        if not csv_filename:
            return "Error: CSV filename is required."

        if not csv_filename.endswith('.csv'):
            return "Error: CSV filename must end with .csv extension."

        # Create full path in working directory
        csv_path = self.working_directory / csv_filename
        resolved_csv_path = str(csv_path.resolve())

        if not self._validate_file_path(resolved_csv_path):
            return "Error: Invalid file path."

        try:
            # Get the worksheet
            ws = self.wb[sheet_name]

            # Convert worksheet to list of lists
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))

            # Write to CSV
            import csv

            with open(
                resolved_csv_path, 'w', newline='', encoding='utf-8-sig'
            ) as csvfile:
                writer = csv.writer(csvfile)
                writer.writerows(data)

            return (
                f"Sheet {sheet_name} exported to {csv_filename} "
                f"in {self.working_directory}."
            )
        except Exception as e:
            logger.error(f"Failed to export sheet to CSV: {e}")
            return f"Error: Failed to export sheet {sheet_name} to CSV: {e}"

    def get_rows(
        self,
        sheet_name: str,
        start_row: Optional[int] = None,
        end_row: Optional[int] = None,
    ) -> Union[List[List[Union[str, int, float, None]]], str]:
        r"""Retrieve rows of data from a sheet.

        Use this to read data from a sheet. You can get all rows or specify a
        range. Returns actual data as lists, making it easy to process
        programmatically.

        Args:
            sheet_name (str): Name of the sheet to read from.
            start_row (Optional[int]): First row to read (1-based). If None,
                starts from row 1. (default: :obj:`None`)
            end_row (Optional[int]): Last row to read (1-based). If None,
                reads to the end. (default: :obj:`None`)

        Returns:
            Union[List[List[Union[str, int, float, None]]], str]:
                List of rows (each row is a list of cell values) or error
                message.
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Error: Sheet {sheet_name} does not exist."

        ws = self.wb[sheet_name]
        rows = []

        # Get all rows with data
        for row in ws.iter_rows(
            min_row=start_row, max_row=end_row, values_only=True
        ):
            # Skip completely empty rows
            if any(cell is not None for cell in row):
                rows.append(list(row))

        return rows

    def append_row(
        self,
        sheet_name: str,
        row_data: List[Union[str, int, float, None]],
    ) -> str:
        r"""Add a single row to the end of a sheet.

        Use this to add one row of data to the end of existing content.
        For multiple rows, use multiple calls to this function.

        Args:
            sheet_name (str): Name of the target sheet.
            row_data (List[Union[str, int, float, None]]): Single row of data
                to add.

        Returns:
            str: Success confirmation message or error details.
        """
        if not self.wb:
            return "Error: Workbook not initialized."
        if sheet_name not in self.wb.sheetnames:
            return f"Error: Sheet {sheet_name} does not exist."
        ws = self.wb[sheet_name]
        ws.append(row_data)
        return f"Row appended to sheet {sheet_name} successfully."

    def update_row(
        self,
        sheet_name: str,
        row_number: int,
        row_data: List[Union[str, int, float, None]],
    ) -> str:
        r"""Update a specific row in the sheet.

        Use this to replace all data in a specific row with new values.

        Args:
            sheet_name (str): Name of the sheet to modify.
            row_number (int): The row number to update (1-based, where 1 is
                first row).
            row_data (List[Union[str, int, float, None]]): New data for the
                entire row.

        Returns:
            str: Success confirmation message or error details.
        """
        if not self.wb:
            return "Error: Workbook not initialized."

        if sheet_name not in self.wb.sheetnames:
            return f"Sheet {sheet_name} does not exist."

        ws = self.wb[sheet_name]

        # Clear the existing row first
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_number, column=col_idx).value = None

        # Set new values
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_number, column=col_idx).value = value

        return f"Row {row_number} updated in sheet {sheet_name} successfully."

    def get_tools(self) -> List[FunctionTool]:
        r"""Returns a list of FunctionTool objects representing the functions
        in the toolkit.

        Returns:
            List[FunctionTool]: A list of FunctionTool objects representing
                the functions in the toolkit.
        """
        return [
            # File operations
            FunctionTool(self.extract_excel_content),
            FunctionTool(self.create_workbook),
            FunctionTool(self.save_workbook),
            FunctionTool(self.delete_workbook),
            FunctionTool(self.export_sheet_to_csv),
            # Sheet operations
            FunctionTool(self.create_sheet),
            FunctionTool(self.delete_sheet),
            FunctionTool(self.clear_sheet),
            # Data reading
            FunctionTool(self.get_rows),
            FunctionTool(self.get_cell_value),
            FunctionTool(self.get_column_data),
            FunctionTool(self.get_range_values),
            FunctionTool(self.find_cells),
            # Data writing
            FunctionTool(self.append_row),
            FunctionTool(self.update_row),
            FunctionTool(self.set_cell_value),
            FunctionTool(self.set_range_values),
            # Structure modification
            FunctionTool(self.delete_rows),
            FunctionTool(self.delete_columns),
        ]
