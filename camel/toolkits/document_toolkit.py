# ========= Copyright 2023-2024 @ CAMEL-AI.org. All Rights Reserved. =========
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ========= Copyright 2023-2024 @ CAMEL-AI.org. All Rights Reserved. =========


from collections import defaultdict
from pathlib import Path
from typing import (
    TYPE_CHECKING,
    Callable,
    Dict,
    List,
    Optional,
    Set,
    Tuple,
)
from urllib.parse import urlparse

from camel.loaders import MarkItDownLoader
from camel.logger import get_logger
from camel.toolkits.base import BaseToolkit
from camel.toolkits.function_tool import FunctionTool
from camel.utils import MCPServer, retry_on_error
from camel.utils.commons import dependencies_required

if TYPE_CHECKING:
    import pandas as pd
    from tabulate import tabulate

logger = get_logger(__name__)

_TEXT_EXTS: Set[str] = {".txt", ".md", ".rtf"}
_DOC_EXTS: Set[str] = {".pdf", ".doc", ".docx", ".ppt", ".pptx", ".odt"}
_EXCEL_EXTS: Set[str] = {".xls", ".xlsx", ".csv"}
_ARCHIVE_EXTS: Set[str] = {".zip"}
_WEB_EXTS: Set[str] = {".html", ".htm", ".xml"}
_CODE_EXTS: Set[str] = {".py", ".js", ".java", ".cpp", ".c", ".go", ".rs"}
_DATA_EXTS: Set[str] = {".json", ".jsonl", ".jsonld", ".xml", ".yaml", ".yml"}


class FileEditResult:
    def __init__(
        self,
        success: bool,
        message: str,
        old_content: str = "",
        new_content: str = "",
    ):
        self.success = success
        self.message = message
        self.old_content = old_content
        self.new_content = new_content
        self.changes_lines = self.new_content.count('\n')


class _LoaderWrapper:
    r"""Uniform interface wrapper for different document loaders.

    Every loader exposes parse_file(path) -> str method for consistent usage.

    Args:
        loader (object): The underlying loader instance.
    """

    def __init__(self, loader: object):
        r"""Initialize the loader wrapper."""
        self._loader = loader

    def parse_file(self, path: str) -> Tuple[bool, str]:
        r"""Convert a file to plain text using the wrapped loader.

        Args:
            path (str): Path to the file to be converted.

        Returns:
            Tuple[bool, str]: A tuple containing success status and either
                the extracted plain text content or an error message.
        """
        # Try parse_file method first (for loaders that have it)
        if hasattr(self._loader, "parse_file"):
            try:
                content = self._loader.parse_file(path)  # type: ignore[attr-defined]
                return True, content
            except Exception as e:
                return False, f"Loader failed to convert file {path}: {e}"

        # Try convert_file method (for MarkItDownLoader)
        if hasattr(self._loader, "convert_file"):
            try:
                content = self._loader.convert_file(path)  # type: ignore[attr-defined]
                return True, content
            except Exception as e:
                return False, f"Loader failed to convert file {path}: {e}"

        return (
            False,
            "Loader does not support parse_file or convert_file method",
        )


def _init_loader(loader_kwargs: Optional[dict] | None) -> _LoaderWrapper:
    r"""Initialize a document loader.

    Args:
        loader_kwargs (Optional[dict]): Additional keyword arguments for the
            loader.

    Returns:
        _LoaderWrapper: Wrapped loader instance.

    Raises:
        ImportError: If the required loader is not available.
    """
    loader_kwargs = loader_kwargs or {}

    if MarkItDownLoader is None:
        raise ImportError("MarkItDownLoader unavailable.")
    return _LoaderWrapper(MarkItDownLoader(**loader_kwargs))


@MCPServer()
class DocumentToolkit(BaseToolkit):
    r"""A comprehensive toolkit for processing various document formats.

    This toolkit can extract plain-text content from local files, remote URLs,
    ZIP archives, and webpages. It supports Office documents, PDFs,
    Excel files, code files, and more.
    """

    def __init__(
        self,
        *,
        cache_dir: str | Path | None = None,
        timeout: Optional[float] = None,
        loader_kwargs: Optional[dict] = None,
        enable_cache: bool = True,
    ) -> None:
        r"""Initialize the DocumentToolkit.

        Args:
            cache_dir (str | Path | None): Directory for caching processed
                documents. Defaults to ~/.cache/camel/documents.
            timeout (Optional[float]): The timeout for the toolkit.
                (default: :obj:`None`)
            loader_kwargs (Optional[dict]): Additional arguments for the
                primary loader.
            enable_cache (bool): Whether to enable disk and memory caching.
        """
        super().__init__(timeout=timeout)
        self._file_history: Dict[str, List[FileEditResult]] = defaultdict(list)

        self.enable_cache = enable_cache
        if self.enable_cache:
            self.cache_dir: Optional[Path] = Path(
                cache_dir or "~/.cache/camel/documents"
            ).expanduser()
            self.cache_dir.mkdir(parents=True, exist_ok=True)
        else:
            self.cache_dir = None

        # Initialize primary document loader
        self._loader = _init_loader(loader_kwargs)
        logger.info("DocumentProcessingToolkit initialised")

        # Initialize in-memory cache (keyed by SHA-256 of path + mtime)
        self._cache: Dict[str, str] = {}

    def insert(self, path: str, insert_line: int, new_str: str):
        r"""Insert a new string at a specific line in the file.
        Args:
            path (str): The path of the file to be inserted.
            insert_line (int): The line number to insert the new string.
            new_str (str): The new string to be inserted.
        Returns:
            str: A message indicating the result of the operation.
        """
        try:
            # Get file extension first
            if "." in path:
                suffix = f".{path.lower().rsplit('.', 1)[-1]}"
            else:
                suffix = ""
            allowed_exts = _TEXT_EXTS | _CODE_EXTS | _DATA_EXTS
            if suffix not in allowed_exts:
                return (
                    f"Unsupported file type: {suffix} supported file types "
                    f"are {_TEXT_EXTS} {_CODE_EXTS} {_DATA_EXTS}"
                )

            file_text = self.read_file_content(path).expandtabs()
            new_str = new_str.expandtabs()
            file_text_lines = file_text.split('\n')
            n_lines_file = len(file_text_lines)

            if insert_line < 0 or insert_line > n_lines_file:
                return (
                    f"Invalid line number: {insert_line}. "
                    f"The file has {n_lines_file} lines."
                )

            new_str_lines = new_str.split('\n')
            new_file_text_lines = (
                file_text_lines[:insert_line]
                + new_str_lines
                + file_text_lines[insert_line:]
            )
            new_file_text = '\n'.join(new_file_text_lines)

            # Try to write the file
            write_result = self.write_file(path, new_file_text)

            # Check if write was successful
            if write_result.startswith("Error"):
                return write_result

            success_msg = (
                f"Successfully inserted {new_str} at line "
                f"{insert_line} in {path}."
            )

            self._file_history[path].append(
                FileEditResult(
                    success=True,
                    message=success_msg,
                    old_content=file_text,
                    new_content=new_file_text,
                )
            )
            return success_msg

        except Exception as e:
            error_msg = (
                f"Error inserting content at line {insert_line} in {path}: {e}"
            )
            self._file_history[path].append(
                FileEditResult(
                    success=False,
                    message=error_msg,
                )
            )
            return error_msg

    def undo_edit(self, path: str):
        r"""Undo the last edit operation on the file.
        Args:
            path (str): The path of the file to be undone.
        Returns:
            str: A message indicating the result of the operation.
        """
        if path not in self._file_history:
            return f"No edit history found for {path}."
        last_edit = self._file_history[path][-1]
        self.write_file(path, last_edit.old_content)
        self._file_history[path].pop()
        return f"Successfully undone the last edit in {path}."

    def read_file_content(
        self, path: str, start_line: int = 0, end_line: Optional[int] = None
    ) -> str:
        r"""Read the content of a file.

        Args:
            path (str): The path of the file to be read.
            start_line (int): Start line number.
            end_line (Optional[int]): End line number.

        Returns:
            str: The content of the file.
        """
        try:
            cache_key = self._hash_key(path)
            if self.enable_cache and cache_key in self._cache:
                logger.debug("Cache hit: %s", path)
                return self._cache[cache_key]

            # Get file extension first
            suffix = (
                f".{path.lower().rsplit('.', 1)[-1]}" if "." in path else ""
            )
            allowed_exts = _TEXT_EXTS | _CODE_EXTS | _DATA_EXTS
            if suffix not in allowed_exts:
                return (
                    f"Unsupported file type: {suffix} supported file types "
                    f"are {_TEXT_EXTS} {_CODE_EXTS} {_DATA_EXTS}"
                )
            if suffix in _TEXT_EXTS:
                ok, content = self._handle_text_file(
                    path, start_line, end_line
                )
            elif suffix in _CODE_EXTS:
                ok, content = self._handle_code_file(
                    path, start_line, end_line
                )
            elif suffix in _DATA_EXTS:
                ok, content = self._handle_data_file(
                    path, start_line, end_line
                )
            else:
                return (
                    f"Unsupported file type: {suffix} supported file types "
                    f"are {_TEXT_EXTS} {_CODE_EXTS} {_DATA_EXTS}"
                )
            if ok:
                return content
            else:
                return f"Error reading file {path}: {content}"
        except Exception as e:
            return f"Error reading file {path}: {e}"

    def replace_file_content(
        self,
        path: str,
        old_content: str,
        new_content: str,
    ) -> str:
        r"""Replace specified string in the file.
        Use for updating specific content in the file.

        Args:
            path (str): The path of the file to be replaced.
            old_content (str): The old content to be replaced.
            new_content (str): The new content to be replaced.

        Returns:
            str: A message indicating the result of the operation.
        """
        try:
            file_path = Path(path).expanduser()
            file_text = file_path.read_text(encoding="utf-8")
            old_content = old_content.expandtabs()
            new_content = new_content.expandtabs()

            occurrences = file_text.count(old_content)
            if occurrences == 0:
                return (
                    f"No replacement performed: '{old_content}' not found in "
                    f"{path}."
                )
            elif occurrences > 1:
                file_text_line = file_text.split('\n')
                lines = [
                    idx + 1
                    for idx, line in enumerate(file_text_line)
                    if old_content in line
                ]
                return (
                    f"Multiple occurrences of '{old_content}' found in {path} "
                    f"at lines {lines}. Please specify the line number to "
                    f"replace."
                )

            new_file_content = file_text.replace(old_content, new_content)

            file_path.write_text(new_file_content, encoding="utf-8")
            self._file_history[path].append(
                FileEditResult(
                    success=True,
                    message=f"Successfully replaced content in {path}.",
                    old_content=file_text,
                    new_content=new_file_content,
                )
            )
            return (
                f"Successfully replaced content in {path}. Review the changes "
                f"and make sure they are as expected. Edit the file again if "
                f"necessary."
            )
        except Exception as e:
            return f"Error replacing content in {path}: {e}"

    def write_file(self, path: str, file: str) -> str:
        r"""
        Create or overwrite a file at the given path with the provided content.

        Args:
            path (str): The path of the file to write.
            file (str): The content to write to the file.

        Returns:
            str: A message indicating the result of the operation.
        """
        try:
            file_path = Path(path).expanduser()
            file_path.write_text(file, encoding="utf-8")
            return f"File written at {path}."
        except Exception as e:
            return f"Error writing file {path}: {e}"

    def find_file(
        self,
        pattern: str,
        directory: Optional[str] = None,  # <-- change from Optional[Path]
        recursive: bool = False,
        include_hidden: bool = False,
    ) -> Optional[list[str]]:
        """
        Find files in the specified directory that match the given pattern.

        Args:
            pattern (str): The pattern to match.
            directory (Optional[str]): The directory to search.
            recursive (bool): Whether to search recursively.
            include_hidden (bool): Whether to include hidden files.

        Returns:
            List[str]: A list of file paths that match the pattern.
        """
        import glob

        if directory is None:
            search_dir = Path.cwd()
        else:
            search_dir = Path(directory).resolve()

        if not search_dir.exists():
            raise FileNotFoundError(f"Directory not found: {search_dir}")

        if not search_dir.is_dir():
            raise ValueError(f"Path is not a directory: {search_dir}")

        try:
            matches = []

            if recursive:
                # Use ** for recursive search
                if '/' in pattern or '\\' in pattern:
                    # Pattern includes path separators
                    glob_pattern = str(search_dir / pattern)
                else:
                    # Simple filename pattern
                    glob_pattern = str(search_dir / "**" / pattern)

                file_paths = glob.glob(glob_pattern, recursive=True)
            else:
                # Search only in the specified directory
                glob_pattern = str(search_dir / pattern)
                file_paths = glob.glob(glob_pattern, recursive=False)

            # Filter results
            for file_path in file_paths:
                path_obj = Path(file_path)

                # Skip directories (only return files)
                if not path_obj.is_file():
                    continue

                # Handle hidden files
                if not include_hidden and path_obj.name.startswith('.'):
                    continue

                matches.append(str(path_obj))

            # Sort results for consistent output
            matches.sort()

            logger.debug(
                f"Found {len(matches)} files matching pattern '{pattern}' in "
                f"{search_dir}"
            )
            return matches
        except Exception as e:
            logger.error(f"Error finding files: {e}")
            return None

    def _resolve_local_path(self, path: str) -> str:
        r"""If path is a URL (not a web page), download it and return the
        local path. Otherwise, return the original path.

        Args:
            path (str): File path or URL to resolve. Can be a local file path,
                web page URL, or downloadable file URL.

        Returns:
            str: Local file path. If the input was a downloadable file URL,
                returns the path to the downloaded file. If the input was a
                local path or web page URL, returns the original path
                unchanged.
        """
        suffix = f".{path.lower().rsplit('.', 1)[-1]}" if "." in path else ""
        is_url = self._is_url(path)
        if is_url and suffix and suffix not in _WEB_EXTS:
            local_path = self._download_file(path)
            return str(local_path)
        return path

    # Public API methods
    @retry_on_error()
    def extract_document_content(self, document_path: str) -> Tuple[bool, str]:
        r"""Extract the content of a given document (or URL) and return the
        processed text.

        It may filter out some information, resulting in inaccurate content.

        Args:
            document_path (str): The path of the document to be processed,
                either a local path or a URL. It can process Office documents,
                PDFs, Excel files, code files, zip files, and webpages, etc.

        Returns:
            Tuple[bool, str]: A tuple containing a boolean indicating whether
                the document was processed successfully, and the content of the
                document (if success).
        """
        import traceback

        try:
            cache_key = self._hash_key(document_path)
            if self.enable_cache and cache_key in self._cache:
                logger.debug("Cache hit: %s", document_path)
                return True, self._cache[cache_key]

            # Use the new helper to resolve local path if needed
            resolved_path = self._resolve_local_path(document_path)
            suffix = (
                f".{resolved_path.lower().rsplit('.', 1)[-1]}"
                if "." in resolved_path
                else ""
            )

            # Process based on file type
            ok, txt = self._process_by_file_type(resolved_path, suffix)
            return self._cache_and_return(cache_key, ok, txt)

        except Exception as exc:
            logger.error("Failed to extract %s: %s", document_path, exc)
            logger.debug(traceback.format_exc())
            return False, f"Error extracting document {document_path}: {exc}"

    def _process_by_file_type(
        self, document_path: str, suffix: str
    ) -> Tuple[bool, str]:
        r"""Process document based on its file type.

        Args:
            document_path (str): Path to the document
            suffix (str): File extension

        Returns:
            Tuple[bool, str]: Success status and extracted content
        """
        # Define handler mapping for different file types
        handlers: List[Tuple[set, Callable[[str], Tuple[bool, str]]]] = [
            (_EXCEL_EXTS, self._handle_excel),
            (_ARCHIVE_EXTS, self._handle_zip),
            (_DATA_EXTS, self._handle_data_file),
            (_CODE_EXTS, self._handle_code_file),
            (_TEXT_EXTS, self._handle_text_file),
            (_WEB_EXTS, self._handle_html_file),
        ]

        # Try each handler based on file extension
        for extensions, handler in handlers:
            if suffix in extensions:
                return handler(document_path)

        # Handle Office/PDF documents
        if suffix in _DOC_EXTS:
            return self._loader.parse_file(document_path)

        # Check if it's a webpage (for URLs without web extensions)
        if self._is_webpage(document_path):
            return self._handle_webpage(document_path)

        # Fallback to generic loader
        return self._loader.parse_file(document_path)

    def _is_url(self, path: str) -> bool:
        r"""Check if the path is a URL.

        Args:
            path (str): String to check for URL format. Can be a file path,
                URL, or any other string.

        Returns:
            bool: True if the path is a valid URL with both scheme (e.g.,
                'http', 'https') and network location (e.g., 'example.com').
                False otherwise.
        """
        parsed = urlparse(path)
        return bool(parsed.scheme and parsed.netloc)

    # Webpage processing methods (ported from original code)
    def _is_webpage(self, url: str) -> bool:
        r"""Determine whether the given URL points to a webpage.

        Args:
            url (str): The URL to check.

        Returns:
            bool: True if the URL is a webpage, False otherwise.
        """
        import mimetypes

        import requests

        try:
            parsed_url = urlparse(url)
            is_url = all([parsed_url.scheme, parsed_url.netloc])
            if not is_url:
                return False

            path = parsed_url.path
            file_type, _ = mimetypes.guess_type(path)
            if file_type is not None and "text/html" in file_type:
                return True

            response = requests.head(url, allow_redirects=True, timeout=10)
            content_type = response.headers.get("Content-Type", "").lower()
            return True if "text/html" in content_type else False

        except requests.exceptions.RequestException as e:
            logger.warning(f"Error while checking the URL: {e}")
            return False
        except TypeError:
            return True

    def _handle_webpage(self, url: str) -> Tuple[bool, str]:
        r"""Extract content from a webpage URL.

        Args:
            url (str): The URL of the webpage to process.

        Returns:
            Tuple[bool, str]: Success status and extracted content.
        """
        try:
            extracted_text = self._extract_webpage_content(url)
            return True, extracted_text
        except Exception as e:
            logger.warning(
                f"FireCrawl failed for {url}: {e}, falling back to MarkItDown"
            )
            try:
                # Try using MarkItDown as fallback
                return self._loader.parse_file(url)
            except Exception as e2:
                logger.error(
                    f"All webpage processing methods failed for {url}: {e2}"
                )
                return (
                    False,
                    f"Failed to extract content from the webpage: {e2}",
                )

    @retry_on_error()
    @dependencies_required('crawl4ai')
    def _extract_webpage_content(self, url: str) -> str:
        r"""Extract webpage content using the Crawl4AI loader.

        Args:
            url (str): The URL of the webpage to extract content from.

        Returns:
            str: The extracted webpage content in markdown format.

        Raises:
            ImportError: If Crawl4AI library is not available.
            Exception: If scraping fails or returns no content.
        """
        from camel.loaders import Crawl4AI

        # Initialize Crawl4AI crawler
        crawler = Crawl4AI()

        try:
            # Use asyncio to run the async scrape method
            import asyncio

            scrape_result = asyncio.run(crawler.scrape(url))

            logger.debug(f"Extracted data from {url}: {scrape_result}")

            # Refactored: Use try/except for extracting markdown content
            try:
                markdown_content = scrape_result['markdown']
                if markdown_content:
                    return str(markdown_content)
                else:
                    return "No content found on the webpage."
            except (KeyError, TypeError):
                return "Error while crawling the webpage."

        except Exception as e:
            logger.error(f"Error scraping {url}: {e!s}")
            return f"Error while crawling the webpage: {e!s}"

    @dependencies_required('tabulate')
    def _convert_to_markdown(self, df: 'pd.DataFrame') -> str:
        r"""Convert DataFrame to Markdown format table.

        Args:
            df (pd.DataFrame): The pandas DataFrame to convert to markdown
                format.

        Returns:
            str: The DataFrame content formatted as a markdown table with pipe
                separators.
        """
        md_table = tabulate(df, headers='keys', tablefmt='pipe')
        return str(md_table)

    def _extract_sheet_cell_info(self, ws):
        r"""Extract cell info from a worksheet.

        Args:
            ws: Worksheet object from openpyxl library containing cells to
                extract information from.

        Returns:
            list: A list of dictionaries, where each dictionary contains
                information about a single cell:
                - index (str): Cell position in format
                  "{row_number}{column_letter}" (e.g., "1A", "2B")
                - value: The cell's value (can be str, int, float, datetime,
                  or None)
                - font_color (str or None): RGB hex color code of the font
                  (e.g., "FF0000") or None if no color
                - fill_color (str or None): RGB hex color code of the cell
                  background or None if no fill color
        """
        cell_info_list = []
        for row in ws.iter_rows():
            for cell in row:
                row_num = cell.row
                col_letter = cell.column_letter
                cell_value = cell.value
                font_color = None
                if (
                    cell.font
                    and cell.font.color
                    and "rgb=None" not in str(cell.font.color)
                ):
                    font_color = cell.font.color.rgb
                fill_color = None
                if (
                    cell.fill
                    and cell.fill.fgColor
                    and "rgb=None" not in str(cell.fill.fgColor)
                ):
                    fill_color = cell.fill.fgColor.rgb
                cell_info_list.append(
                    {
                        "index": f"{row_num}{col_letter}",
                        "value": cell_value,
                        "font_color": font_color,
                        "fill_color": fill_color,
                    }
                )
        return cell_info_list

    @dependencies_required('pandas', 'openpyxl', 'xls2xlsx')
    def _extract_excel_content(self, document_path: str) -> str:
        r"""Extract detailed cell information from an Excel file, including
        multiple sheets.

        Args:
            document_path (str): The path of the Excel file.

        Returns:
            str: Extracted excel information, including details of each sheet.
        """
        import pandas as pd
        from openpyxl import load_workbook
        from xls2xlsx import XLS2XLSX

        logger.debug(
            f"Calling extract_excel_content with document_path: "
            f"{document_path}"
        )

        if not (
            document_path.endswith("xls")
            or document_path.endswith("xlsx")
            or document_path.endswith("csv")
        ):
            logger.error("Only xls, xlsx, csv files are supported.")
            return (
                f"Failed to process file {document_path}: "
                f"It is not excel format. Please try other ways."
            )

        if document_path.endswith("csv"):
            try:
                df = pd.read_csv(document_path)
                md_table = self._convert_to_markdown(df)
                return f"CSV File Processed:\n{md_table}"
            except Exception as e:
                logger.error(f"Failed to process file {document_path}: {e}")
                return f"Failed to process file {document_path}: {e}"

        if document_path.endswith("xls"):
            output_path = document_path.replace(".xls", ".xlsx")
            x2x = XLS2XLSX(document_path)
            x2x.to_xlsx(output_path)
            document_path = output_path

        # Load the Excel workbook
        wb = load_workbook(document_path, data_only=True)
        sheet_info_list = []

        # Iterate through all sheets
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            cell_info_list = self._extract_sheet_cell_info(ws)
            # Convert the sheet to a DataFrame and then to markdown
            sheet_df = pd.read_excel(
                document_path, sheet_name=sheet, engine='openpyxl'
            )
            markdown_content = self._convert_to_markdown(sheet_df)
            # Collect all information for the sheet
            sheet_info = {
                "sheet_name": sheet,
                "cell_info_list": cell_info_list,
                "markdown_content": markdown_content,
            }
            sheet_info_list.append(sheet_info)

        result_str = ""
        for sheet_info in sheet_info_list:
            result_str += (
                f"\n            Sheet Name: {sheet_info['sheet_name']}\n"
                f"            Cell information list:\n"
                f"            {sheet_info['cell_info_list']}\n\n"
                f"            Markdown View of the content:\n"
                f"            {sheet_info['markdown_content']}\n\n"
                f"            {'-' * 40}\n"
            )

        return result_str

    def _handle_excel(self, path: str) -> Tuple[bool, str]:
        r"""Process Excel and CSV files using integrated Excel processing.

        Args:
            path (str): Path to the Excel or CSV file.

        Returns:
            Tuple[bool, str]: Success status and Excel content.
        """
        try:
            res = self._extract_excel_content(path)
            # Check if the result indicates an error
            if res.startswith("Failed to process file"):
                return False, res
            return True, res
        except Exception as e:
            logger.error(f"Error processing Excel file {path}: {e}")
            return False, f"Error processing Excel file {path}: {e}"

    @dependencies_required('xmltodict', 'pyyaml')
    def _handle_data_file(
        self, path: str, start_line: int = 0, end_line: Optional[int] = None
    ) -> Tuple[bool, str]:
        r"""Process structured data files (JSON, XML, YAML).

        Args:
            path (str): Path to the data file.

        Returns:
            Tuple[bool, str]: Success status and parsed data content.
        """
        try:
            # Read the file content (with optional line filtering)
            ok, content = self._read_file_content_with_lines(
                path, start_line, end_line
            )
            if not ok:
                return False, content

            # Process based on file type
            suffix = Path(path).suffix.lower()
            return self._process_data_content(content, suffix, path)

        except Exception as e:
            return False, f"Error processing data file {path}: {e}"

    def _read_file_content_with_lines(
        self, path: str, start_line: int = 0, end_line: Optional[int] = None
    ) -> Tuple[bool, str]:
        r"""Read file content with optional line-based filtering.

        Args:
            path (str): Path to the file.
            start_line (int): Starting line number.
            end_line (Optional[int]): Ending line number.

        Returns:
            Tuple[bool, str]: Success status and file content.
        """
        try:
            with open(path, "r", encoding="utf-8") as f:
                lines = f.readlines()

            total_lines = len(lines)

            # If reading entire file
            if start_line == 0 and end_line is None:
                return True, ''.join(lines)

            # Validate start line
            if start_line >= total_lines:
                return (
                    False,
                    "Start line {} exceeds file length ({} lines)".format(
                        start_line, total_lines
                    ),
                )

            # Determine end line
            if end_line is None:
                end_line = total_lines - 1
            elif end_line >= total_lines:
                end_line = total_lines - 1

            # Extract the specified lines
            selected_lines = lines[start_line : end_line + 1]
            return True, ''.join(selected_lines)

        except Exception as e:
            return False, f"Error reading file {path}: {e}"

    def _process_data_content(
        self, content: str, suffix: str, path: str
    ) -> Tuple[bool, str]:
        r"""Process data content based on file type.

        Args:
            content (str): File content to process.
            suffix (str): File extension.
            path (str): Original file path (for error messages).

        Returns:
            Tuple[bool, str]: Success status and processed content.
        """
        if suffix in [".json", ".jsonl", ".jsonld"]:
            return self._process_json_content(content, suffix)
        elif suffix in [".yaml", ".yml"]:
            return self._process_yaml_content(content)
        elif suffix == ".xml":
            return self._process_xml_content(content)
        else:
            return False, f"Unsupported data file format: {suffix}"

    def _process_json_content(
        self, content: str, suffix: str
    ) -> Tuple[bool, str]:
        r"""Process JSON content based on JSON variant.

        Args:
            content (str): JSON content to parse.
            suffix (str): File extension to determine JSON variant.

        Returns:
            Tuple[bool, str]: Success status and formatted JSON content.
        """
        import json

        try:
            if suffix == ".jsonl":
                return self._process_jsonl_content(content)
            else:
                return self._process_regular_json_content(content)
        except json.JSONDecodeError as e:
            return False, f"Error parsing JSON: {e}"

    def _process_jsonl_content(self, content: str) -> Tuple[bool, str]:
        r"""Process JSON Lines format content.

        Args:
            content (str): JSONL content to parse.

        Returns:
            Tuple[bool, str]: Success status and formatted content.
        """
        import json

        json_objects = []
        for line in content.strip().split('\n'):
            if line.strip():  # Skip empty lines
                try:
                    json_objects.append(json.loads(line.strip()))
                except json.JSONDecodeError as e:
                    return False, f"Error parsing JSON line: {e}"

        formatted_content = json.dumps(
            json_objects, indent=2, ensure_ascii=False
        )
        return True, formatted_content

    def _process_regular_json_content(self, content: str) -> Tuple[bool, str]:
        r"""Process regular JSON format content.

        Args:
            content (str): JSON content to parse.

        Returns:
            Tuple[bool, str]: Success status and formatted content.
        """
        import json

        parsed_content = json.loads(content)

        if isinstance(parsed_content, (list, dict)):
            formatted_content = json.dumps(
                parsed_content, indent=2, ensure_ascii=False
            )
        else:
            formatted_content = str(parsed_content)

        return True, formatted_content

    def _process_yaml_content(self, content: str) -> Tuple[bool, str]:
        r"""Process YAML format content.

        Args:
            content (str): YAML content to parse.

        Returns:
            Tuple[bool, str]: Success status and formatted content.
        """
        import yaml

        try:
            # Handle multiple YAML documents in one file
            documents = list(yaml.safe_load_all(content))
            parsed_content = documents[0] if len(documents) == 1 else documents

            if isinstance(parsed_content, (list, dict)):
                formatted_content = yaml.dump(
                    parsed_content,
                    default_flow_style=False,
                    allow_unicode=True,
                )
            else:
                formatted_content = str(parsed_content)

            return True, formatted_content
        except yaml.YAMLError as e:
            return False, f"Error parsing YAML: {e}"

    def _process_xml_content(self, content: str) -> Tuple[bool, str]:
        r"""Process XML format content.

        Args:
            content (str): XML content to parse.

        Returns:
            Tuple[bool, str]: Success status and formatted content.
        """
        import json

        import xmltodict  # type: ignore[import-untyped]

        try:
            # Try to parse as structured XML first
            data = xmltodict.parse(content)
            logger.debug(f"The extracted xml data is: {data}")

            # Format output nicely
            formatted_content = json.dumps(data, indent=2, ensure_ascii=False)
            return True, formatted_content
        except Exception as e:
            # If parsing fails, return raw content
            logger.debug(f"XML parsing failed, returning raw content: {e}")
            return True, content

    def _handle_code_file(
        self, path: str, start_line: int = 0, end_line: Optional[int] = None
    ) -> Tuple[bool, str]:
        r"""Process source code files by reading them as plain text.

        Args:
            path (str): Path to the code file.

        Returns:
            Tuple[bool, str]: Success status and code content.
        """
        try:
            return self._handle_text_file(path, start_line, end_line)
        except Exception as e:
            return False, f"Error processing code file {path}: {e}"

    def _handle_text_file(
        self, path: str, start_line: int = 0, end_line: Optional[int] = None
    ) -> Tuple[bool, str]:
        r"""Process plain text files by reading them directly.

        Args:
            path (str): Path to the text file.
            start_line (int): Start line number.
            end_line (Optional[int]): End line number.

        Returns:
            Tuple[bool, str]: Success status and text content.
        """
        try:
            file_path = Path(path).expanduser()

            # Read all lines from file
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            total_lines = len(lines)

            # Select lines based on start_line and end_line arguments
            if start_line == 0 and end_line is None:
                # Case 1: No line range specified, return all lines
                selected_lines = lines
            elif end_line is None:
                # Case 2: Only start_line specified,
                # return from start_line to end
                if start_line >= total_lines:
                    return (
                        False,
                        f"Start line {start_line} exceeds file length "
                        f"({total_lines} lines)",
                    )
                selected_lines = lines[start_line:]
            elif start_line == 0:
                # Case 3: Only end_line specified,
                # return from start to end_line (inclusive)
                if end_line >= total_lines:
                    end_line = total_lines - 1
                selected_lines = lines[: end_line + 1]
            else:
                # Case 4: Both start_line and
                # end_line specified,
                # return the range
                if start_line >= total_lines:
                    return (
                        False,
                        f"Start line {start_line} exceeds file length "
                        f"({total_lines} lines)",
                    )
                if end_line >= total_lines:
                    end_line = total_lines - 1
                selected_lines = lines[start_line : end_line + 1]

            # Join lines and return
            text = ''.join(selected_lines)
            return True, text
        except Exception as e:
            return False, f"Error reading text file {path}: {e}"

    def _handle_html_file(
        self, path: str, start_line: int = 0, end_line: Optional[int] = None
    ) -> Tuple[bool, str]:
        r"""Process local HTML/XML files.

        Args:
            path (str): Path to the HTML/XML file.

        Returns:
            Tuple[bool, str]: Success status and processed content.
        """
        # Try using the generic loader first
        ok, txt = self._loader.parse_file(path)
        if ok:
            return True, txt

        # Fallback to raw text reading
        try:
            return self._handle_text_file(path, start_line, end_line)
        except Exception as e2:
            return False, f"Error processing HTML file {path}: {e2}"

    # ZIP archive processing methods
    def _handle_zip(self, path_or_url: str) -> Tuple[bool, str]:
        r"""Process ZIP archive files by extracting and processing all
        contained files.

        Args:
            path_or_url (str): Path or URL to the ZIP file.

        Returns:
            Tuple[bool, str]: Success status and combined content of all files.
        """
        import traceback
        import zipfile

        try:
            if not self.cache_dir:
                return (
                    False,
                    "Cache directory not available for ZIP processing",
                )

            zip_path = self._ensure_local_zip(path_or_url)
            extract_dir = (
                self.cache_dir / f"unzip_{self._short_hash(zip_path)}"
            )
            extract_dir.mkdir(parents=True, exist_ok=True)

            with zipfile.ZipFile(zip_path, "r") as zf:
                zf.extractall(extract_dir)
                # Log the files in the ZIP
                logger.info(f"ZIP contains files: {zf.namelist()}")

            parts: List[str] = []
            failed_files: List[str] = []

            for file in extract_dir.rglob("*"):
                if file.is_file():
                    logger.info(f"Processing file: {file}")
                    ok, text = self.extract_document_content(str(file))
                    if ok:
                        parts.append(f"=== File: {file.name} ===\n{text}")
                    else:
                        failed_files.append(f"{file.name}: {text}")

            result = "\n\n".join(parts) if parts else ""
            if failed_files:
                result += "\n\nFailed to process:\n" + "\n".join(failed_files)

            return True, result
        except Exception as exc:
            logger.error("ZIP processing failed for %s: %s", path_or_url, exc)
            logger.debug(traceback.format_exc())
            return False, f"Error processing ZIP {path_or_url}: {exc}"

    def _ensure_local_zip(self, path_or_url: str) -> Path:
        r"""Ensure ZIP file is available locally, downloading if necessary.

        Args:
            path_or_url (str): Path or URL to the ZIP file.

        Returns:
            Path: Local path to the ZIP file.
        """
        # Use the new helper to resolve local path if needed
        resolved_path = self._resolve_local_path(path_or_url)
        return Path(resolved_path).expanduser().resolve()

    def _download_file(self, url: str) -> Path:
        r"""Download a file from URL to the local cache directory.

        Args:
            url (str): The URL of the file to download.

        Returns:
            Path: Local path to the downloaded file.

        Raises:
            RuntimeError: If cache directory is not available.
            requests.RequestException: If download fails.
            IOError: If file cannot be saved to disk.
        """
        import re

        import requests

        if not self.cache_dir:
            raise RuntimeError(
                "Cache directory not available for downloading files"
            )

        try:
            resp = requests.get(url, stream=True, timeout=60)
            resp.raise_for_status()
        except requests.RequestException as e:
            logger.error("Failed to download %s: %s", url, e)
            raise

        # Try to get filename from Content-Disposition header
        cd = resp.headers.get("Content-Disposition", "")
        filename = None
        if "filename" in cd:
            match = re.findall(r'filename[*]?=["\']?([^"\';\r\n]+)', cd)
            if match:
                filename = match[0].strip()

        # Fallback: get filename from URL
        if not filename:
            parsed_url = urlparse(url)
            filename = Path(parsed_url.path).name or "downloaded_file"

            # If no extension, guess from content type
            if not Path(filename).suffix:
                content_type = (
                    resp.headers.get("Content-Type", "").split(';')[0].strip()
                )
                ext_map = {
                    "application/pdf": ".pdf",
                    "application/zip": ".zip",
                    "application/json": ".json",
                    "text/html": ".html",
                    "application/msword": ".doc",
                    "application/vnd.openxmlformats-"
                    "officedocument.wordprocessingml.document": ".docx",
                    "application/vnd.ms-excel": ".xls",
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet": ".xlsx",
                    "text/plain": ".txt",
                    "text/csv": ".csv",
                    "application/x-yaml": ".yaml",
                    "text/yaml": ".yaml",
                    "text/x-yaml": ".yaml",
                }
                ext = ext_map.get(content_type, "")
                filename += ext

        # Sanitize filename for filesystem
        filename = re.sub(r'[<>:"/\\|?*]', '_', filename)

        local_path = self.cache_dir / filename
        logger.info("Downloading %s → %s", url, local_path)

        try:
            with open(local_path, "wb") as f:
                for chunk in resp.iter_content(chunk_size=8192):
                    if chunk:  # Filter out keep-alive chunks
                        f.write(chunk)
        except IOError as e:
            logger.error("Failed to save file %s: %s", local_path, e)
            raise

        return local_path

    # Cache management methods
    def _cache_and_return(
        self, key: str, ok: bool, value: str
    ) -> Tuple[bool, str]:
        r"""Store successful results in cache and return the result tuple.

        Args:
            key (str): Cache key for storing the result.
            ok (bool): Success status of the operation.
            value (str): The processed content.

        Returns:
            Tuple[bool, str]: The input success status and value.
        """
        if ok and self.enable_cache and self.cache_dir:
            self._cache[key] = value
            try:
                (self.cache_dir / f"{key}.txt").write_text(
                    value, encoding="utf-8"
                )
            except Exception as e:  # pragma: no cover
                logger.debug("Cache write failed: %s", e)
        return ok, value

    # Hash generation utilities
    def _hash_key(self, path: str) -> str:
        r"""Generate a stable cache key using SHA-256 hash of path and
        modification time.

        Args:
            path (str): File path or URL to generate key for.

        Returns:
            str: SHA-256 hash string to use as cache key.
        """
        import hashlib

        if self._is_webpage(path):
            # For URLs, use the URL itself as the key component
            return hashlib.sha256(f"{path}:url".encode()).hexdigest()

        p = Path(path)
        mtime = p.stat().st_mtime if p.exists() else 0
        return hashlib.sha256(f"{path}:{mtime}".encode()).hexdigest()

    def _short_hash(self, path: Path) -> str:
        r"""Generate a short hash for directory naming.

        Args:
            path (Path): Path to generate hash for.

        Returns:
            str: Short (8-character) hash string.
        """
        import hashlib

        return hashlib.sha256(str(path).encode()).hexdigest()[:8]

    def get_tools(self) -> List[FunctionTool]:
        r"""Returns a list of FunctionTool objects representing the functions
        in the toolkit.

        Returns:
            List[FunctionTool]: A list of FunctionTool objects representing
                the functions in the toolkit.
        """
        return [
            FunctionTool(self.extract_document_content),
            FunctionTool(self.insert),
            FunctionTool(self.write_file),
            FunctionTool(self.replace_file_content),
            FunctionTool(self.undo_edit),
            FunctionTool(self.find_file),
            FunctionTool(self.read_file_content),
        ]

    async def _extract_async(self, document_path: str) -> str:
        r"""Asynchronous wrapper for document extraction (rarely used).

        Args:
            document_path (str): Path to the document to extract.

        Returns:
            str: Extracted document content.
        """
        ok, content = self._loader.parse_file(document_path)
        if ok:
            return content
        else:
            raise RuntimeError(content)

    def extract_document_content_sync(
        self, document_path: str
    ) -> Tuple[bool, str]:
        r"""Synchronous wrapper for environments without asyncio support.

        Args:
            document_path (str): Path to the document to extract.

        Returns:
            Tuple[bool, str]: Success status and extracted content.
        """
        return self.extract_document_content(document_path)
